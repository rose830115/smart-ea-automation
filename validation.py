"""Pre-flight checks for Smart EA pipeline inputs.

Validates vendor Excel + reference template against the agreed specs before
running the pipeline. Reports violations as a list of human-readable messages
so the web UI can red-flag them and block the run button.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from sheet_names import missing_zones, resolve_zone_sheets

ALLOWED_ZONES: tuple[str, ...] = ("RMW", "PL", "FGW")

# cp 編號允許格式：純數字、數字+小寫字母、或 O 開頭+數字
CP_PATTERN = re.compile(r"^(?:O\d+|\d+[a-z]?)$", re.IGNORECASE)

# area name 允許格式：純數字、或 "Area N"
AREA_NAME_PURE_NUM = re.compile(r"^\d+$")
AREA_NAME_STANDARD = re.compile(r"^Area\s+\d+$")


def _row_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_vendor_file(vendor_path: Path) -> list[str]:
    """Check vendor xlsx against spec A (sheet names, cp format)."""
    violations: list[str] = []
    wb = load_workbook(vendor_path, data_only=True, read_only=True)

    # A1: 每個 zone 至少要有一個分頁（前綴比對，容忍大小寫 / 空白 / 廠區後綴）
    for canonical in missing_zones(wb.sheetnames):
        violations.append(
            f"業務檔找不到「{canonical}」的分頁。"
            f"分頁名開頭需為「{canonical}」，後面可加廠區 / 製程（如 -CKL1、-CKL2 (Cutting）"
        )

    # A2: Col A cp format（對所有實際對應到 zone 的分頁檢查，含同 zone 多頁）
    for sheet, _zone in resolve_zone_sheets(wb.sheetnames):
        ws = wb[sheet]
        for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row:
                continue
            cp = row[0] if len(row) > 0 else None
            cp_str = _row_text(cp)
            if not cp_str:
                continue
            # int 轉成 "3" 這種會被當有效（CP_PATTERN match 純數字）
            if isinstance(cp, float) and cp.is_integer():
                cp_str = str(int(cp))
            if not CP_PATTERN.fullmatch(cp_str):
                violations.append(
                    f"業務檔 sheet「{sheet}」第 {idx} 列 cp 欄位異常：「{cp_str}」"
                    f"（應為議定編號，如 `3b`、`O1`，不可填功能名稱或合併區）"
                )
    return violations


def validate_target_file(target_path: Path) -> list[str]:
    """Check reference template against spec B (Areas sheet area name / zone)."""
    violations: list[str] = []
    wb = load_workbook(target_path, data_only=True, read_only=True)

    if "Areas" not in wb.sheetnames:
        violations.append("公式範本缺少「Areas」sheet")
        return violations

    ws = wb["Areas"]
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        area = row[3] if len(row) > 3 else None
        zone = row[5] if len(row) > 5 else None

        # B1: Col D area name 格式
        area_str = _row_text(area)
        if isinstance(area, float) and area.is_integer():
            area_str = str(int(area))
        if area_str:
            is_pure_num = bool(AREA_NAME_PURE_NUM.fullmatch(area_str))
            is_standard = bool(AREA_NAME_STANDARD.fullmatch(area_str))
            if not (is_pure_num or is_standard):
                violations.append(
                    f"公式範本 Areas 第 {idx} 列 area name 違規：「{area_str}」"
                    f"（只允許純數字或 `Area N` 連號，不可含文字或特殊字元）"
                )

        # B2: Col F zone 格式
        zone_str = _row_text(zone)
        if zone_str and zone_str not in ALLOWED_ZONES:
            violations.append(
                f"公式範本 Areas 第 {idx} 列 zone 違規：「{zone_str}」"
                f"（只接受 RMW / PL / FGW）"
            )
        if area_str and not zone_str:
            violations.append(
                f"公式範本 Areas 第 {idx} 列 area「{area_str}」缺 zone 欄位"
            )
    return violations


def validate_inputs(vendor_path: Path | None, target_path: Path | None) -> list[str]:
    """Run all pre-flight checks. Returns empty list if all pass."""
    violations: list[str] = []
    if vendor_path is not None and Path(vendor_path).exists():
        try:
            violations.extend(validate_vendor_file(Path(vendor_path)))
        except Exception as exc:
            violations.append(f"業務檔讀取失敗：{exc}")
    if target_path is not None and Path(target_path).exists():
        try:
            violations.extend(validate_target_file(Path(target_path)))
        except Exception as exc:
            violations.append(f"公式範本讀取失敗：{exc}")
    return violations
