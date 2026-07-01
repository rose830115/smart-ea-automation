#!/usr/bin/env python3
"""
Smart EA environmental audit automation.

This script converts a vendor field-data workbook into the standardized
Env. Data / Moisture / CFU workbook shape, validates it against a reference
manual workbook when available, and drafts environmental comments.
"""

from __future__ import annotations

import argparse
import copy
import json
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from yims_payload_builder import write_yims_fill_plan
from sheet_names import outside_sheets, resolve_zone_sheets


BASE_NAS = Path("/Volumes/實驗室共用區MRC/#YCT資料區#/4.Smart EA")
DEFAULT_VENDOR = BASE_NAS / "案件資料區/2025/PD#EA-2511R06 1112 ID NB PYS/PYS 112025 Record_of_environment_data_and_swab_sampling_for_audit_team_ver_5.xlsx"
DEFAULT_TARGET = BASE_NAS / "案件資料區/2025/PD#EA-2511R06 1112 ID NB PYS/PYS CFU and Environmental Raw Data (v2.0).xlsx"
DEFAULT_OUTDIR = Path(__file__).resolve().parent / "outputs"

ZONE_LABELS = {
    "RMW": "Raw Material Warehouse",
    "PL": "Production Line",
    "FGW": "Finished Goods Warehouse",
    "OUTSIDE": "Outside",
}

CLASS_LABELS = {
    "a.原材料": "raw materials",
    "b.半成品": "semi-finished goods",
    "c.成品": "finished goods",
    "e.外包裝": "outer packaging",
    "f.設備": "equipment",
    "g.建築本體": "building structures",
    "h.工作人員": "workers",
    "j.生產設備": "production equipment",
    "k.通風與除濕設備": "ventilation and dehumidification equipment",
    "l.其他": "other objects",
}

BACKEND_CLASS_LABELS = {
    "a.原材料": "原材料",
    "b.半成品": "半成品",
    "c.成品": "成品",
    "d.包裝材": "包裝材",
    "e.外包裝": "外包裝",
    "f.設備": "設備",
    "g.建築本體": "建築本體",
    "h.工作人員": "工作人員",
    "i.乘載器具": "乘載器具",
    "j.生產設備": "生產設備",
    "k.通風與除濕設備": "通風與除濕設備",
    "l.其他": "其它",
}

ENV_BACKEND_FIELDS = {
    "Temp. (°C)": "temperature_c",
    "Humidity (%)": "humidity_percent",
    "CO2 (ppm)": "co2_ppm",
    "Wind Flow (m/s)": "wind_flow_m_s",
    "PM10 (μg/m³)": "pm10_ug_m3",
}

MOISTURE_HEADERS = [
    "Checking Point",
    "Report Area",
    "Main Zone",
    "Classification",
    "Object",
    "Moisture (%) - 1",
    "Moisture (%) - 2",
    "Moisture (%) - 3",
    "Moisture (%) - 4",
    "Moisture (%) - 5",
]

ENV_HEADERS = [
    "Checking Point",
    "Report Area",
    "Main Zone",
    "Temp. (°C)",
    "Humidity (%)",
    "CO2 (ppm)",
    "Wind Flow (m/s)",
    "PM10 (μg/m³)",
]

CFU_HEADERS = [
    "Checking Point",
    "Swab",
    "Report Area",
    "Main Zone",
    "Classification",
    "Object",
    "Count",
    "CFU/m²",
]


@dataclass
class ReferenceData:
    area_by_cp: dict[str, str]
    area_info: dict[str, dict[str, str]]
    moisture_class_by_object: dict[str, str]
    cfu_class_by_object: dict[str, str]
    count_by_swab: dict[str, float]
    mold_species_by_swab: dict[str, list[str]]
    target_tables: dict[str, list[dict[str, Any]]]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def object_key(value: Any) -> str:
    return clean_text(value).casefold()


def cp_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = clean_text(value)
    if re.fullmatch(r"\d+\.0+", text):
        return str(int(float(text)))
    return text.upper()


def display_cp(value: Any) -> Any:
    key = cp_key(value)
    if re.fullmatch(r"\d+", key):
        return int(key)
    return key


# 儀器讀數下限: 低於下限視為儀器讀不到, 統一補成下限值
MOISTURE_FLOOR = 6
WIND_FLOW_FLOOR = 0.1


def clamp_floor(value: float | int | None, floor: float) -> float | int | None:
    if value is None:
        return None
    return floor if value < floor else value


def normalize_area_name(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+", text):
        return f"Area {int(text)}"
    return text


def normalize_swab_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = clean_text(value).casefold()
    if not text:
        return ""
    for prefix in ("swab", "sw", "s", "#"):
        if text.startswith(prefix):
            rest = text[len(prefix):].lstrip(" -_")
            if re.fullmatch(r"\d+", rest):
                return rest
            if re.fullmatch(r"[a-z]\d+", rest):
                return rest
            break
    if re.fullmatch(r"\d+", text):
        return text
    if re.fullmatch(r"[a-z]\d+", text):
        return text
    return text


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    text = re.sub(r"\.$", "", text)
    try:
        return float(text)
    except ValueError:
        return None


def maybe_int(value: float | None) -> float | int | None:
    if value is None:
        return None
    if float(value).is_integer():
        return int(value)
    return value


def format_number(value: float | None, digits: int = 1) -> str:
    if value is None:
        return "N/A"
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    return f"{value:.{digits}f}".rstrip("0").rstrip(".")


def join_list(items: list[str]) -> str:
    items = [item for item in items if item]
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return ", ".join(items[:-1]) + f", and {items[-1]}"


def read_table(ws, headers: list[str], min_row: int = 2) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        if not any(v is not None for v in row[: len(headers)]):
            continue
        rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return rows


def read_reference(target_path: Path | None) -> ReferenceData:
    area_by_cp: dict[str, str] = {}
    area_info: dict[str, dict[str, str]] = {}
    moisture_class_by_object: dict[str, str] = {}
    cfu_class_by_object: dict[str, str] = {}
    count_by_swab: dict[str, float] = {}
    mold_species_by_swab: dict[str, list[str]] = defaultdict(list)
    target_tables: dict[str, list[dict[str, Any]]] = {}

    if not target_path or not target_path.exists():
        return ReferenceData(area_by_cp, area_info, moisture_class_by_object, cfu_class_by_object, count_by_swab, mold_species_by_swab, target_tables)

    wb = load_workbook(target_path, data_only=True, read_only=True)

    if "Areas" in wb.sheetnames:
        ws = wb["Areas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            cp, report_area = row[0], row[1]
            if cp is not None and report_area is not None:
                area_by_cp[cp_key(cp)] = normalize_area_name(report_area)
            area, function, zone = row[3], row[4], row[5]
            if area is not None and zone is not None:
                area_info[normalize_area_name(area)] = {
                    "function": clean_text(function),
                    "zone": clean_text(zone),
                }

    if "ITEMLIST" in wb.sheetnames:
        ws = wb["ITEMLIST"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            obj, moisture_cls, cfu_cls = row[0], row[1], row[2]
            if obj:
                if moisture_cls:
                    moisture_class_by_object[object_key(obj)] = clean_text(moisture_cls)
                if cfu_cls:
                    cfu_class_by_object[object_key(obj)] = clean_text(cfu_cls)

    if "Moisture" in wb.sheetnames:
        target_tables["Moisture"] = read_table(wb["Moisture"], MOISTURE_HEADERS)
        for row in target_tables["Moisture"]:
            obj, cls = row["Object"], row["Classification"]
            if obj and cls:
                moisture_class_by_object[object_key(obj)] = clean_text(cls)

    if "CFU" in wb.sheetnames:
        target_tables["CFU"] = read_table(wb["CFU"], CFU_HEADERS)
        for row in target_tables["CFU"]:
            obj, cls = row["Object"], row["Classification"]
            if obj and cls:
                cfu_class_by_object[object_key(obj)] = clean_text(cls)
            swab, count = row.get("Swab"), safe_float(row.get("Count"))
            if swab is not None and count is not None:
                count_by_swab[normalize_swab_key(swab)] = count

    if "Env. Data" in wb.sheetnames:
        target_tables["Env. Data"] = read_table(wb["Env. Data"], ENV_HEADERS)

    if "MISC." in wb.sheetnames:
        ws = wb["MISC."]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if len(row) < 102:
                continue
            swab, count = row[100], safe_float(row[101])
            if swab is not None and count is not None:
                count_by_swab[normalize_swab_key(swab)] = count

    if "菌落計數表" in wb.sheetnames:
        ws = wb["菌落計數表"]
        for col in range(2, ws.max_column + 1):
            swab = ws.cell(2, col).value
            if swab is None:
                continue
            swab_key = normalize_swab_key(swab)
            total = 0.0
            for row in range(3, ws.max_row + 1):
                species = clean_text(ws.cell(row, 1).value)
                if not species or "total cfu" in species.casefold() or "風險菌" in species:
                    continue
                value = safe_float(ws.cell(row, col).value)
                if value is not None:
                    total += value
                    if value > 0 and species not in mold_species_by_swab[swab_key]:
                        mold_species_by_swab[swab_key].append(species)
            if swab_key not in count_by_swab:
                count_by_swab[swab_key] = total if total > 0 else 0.01

    return ReferenceData(area_by_cp, area_info, moisture_class_by_object, cfu_class_by_object, count_by_swab, mold_species_by_swab, target_tables)


def normalize_moisture_object(item: Any, specific: Any) -> str:
    base = clean_text(item)
    detail = clean_text(specific)
    if not base:
        return ""
    lower = base.casefold()
    if detail.casefold() == "wooden" and lower in {"rack", "pallet"}:
        return f"Wooden {base[:1].upper()}{base[1:]}"
    return base


def normalize_swab_object(item: Any, specific: Any) -> str:
    return clean_text(item)


def classify_object(obj: str, kind: str, ref: ReferenceData) -> str:
    key = object_key(obj)
    mapping = ref.moisture_class_by_object if kind == "moisture" else ref.cfu_class_by_object
    if key in mapping:
        return mapping[key]

    if any(x in key for x in ["carton", "inner box", "outer box", "box"]):
        return "e.外包裝"
    if any(x in key for x in ["finished goods", "finished shoe"]):
        return "c.成品"
    if any(x in key for x in ["bonded fabric", "upper", "outsole", "tongue", "semi"]):
        return "b.半成品"
    if any(x in key for x in ["fabric", "leather", "foam", "insole", "mesh", "canvas", "rubber", "suede"]):
        return "a.原材料"
    if any(x in key for x in ["glove", "staff", "operator", "worker", "hand"]):
        return "h.工作人員"
    if any(x in key for x in ["building", "wall", "floor", "ceiling", "door"]):
        return "g.建築本體"
    if any(x in key for x in ["dehumidifier", "filter", "fan", "ventilation", "air conditioner"]):
        return "k.通風與除濕設備"
    if kind == "cfu" and any(x in key for x in ["rack", "table", "template", "machine"]):
        return "j.生產設備"
    if any(x in key for x in ["rack", "pallet", "trolley", "chiller", "organizer", "equipment"]):
        return "f.設備"
    return "l.其他"


def report_area_for_cp(cp: str, ref: ReferenceData, fallback_zone: str) -> str:
    key = cp_key(cp)
    if key in ref.area_by_cp:
        return ref.area_by_cp[key]
    if fallback_zone == "OUTSIDE":
        return "Outside"
    return ""


def main_zone_for(cp: str, report_area: str, ref: ReferenceData, fallback_zone: str) -> str:
    if report_area in ref.area_info:
        return ref.area_info[report_area]["zone"]
    return fallback_zone


def parse_outside_cp(value: Any) -> str:
    text = clean_text(value)
    match = re.search(r"\((O\d+)\)", text, flags=re.IGNORECASE)
    if match:
        return match.group(1).upper()
    match = re.search(r"\bO\d+\b", text, flags=re.IGNORECASE)
    if match:
        return match.group(0).upper()
    return text


def read_vendor(vendor_path: Path, ref: ReferenceData) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(vendor_path, data_only=True, read_only=True)
    env_rows: list[dict[str, Any]] = []
    moisture_rows: list[dict[str, Any]] = []
    cfu_rows: list[dict[str, Any]] = []

    for sheet_name, fallback_zone in resolve_zone_sheets(wb.sheetnames):
        ws = wb[sheet_name]
        current_cp = ""
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not any(v is not None for v in row[:18]):
                continue
            cp_value = row[0]
            if cp_value is not None:
                current_cp = cp_key(cp_value)
            if not current_cp:
                continue

            # 業務檔常見「CP 只寫一次, 同 CP 下多列環境量測」, 每列只要有
            # 任一環境數值就 append, 後台會自動取平均
            env_values = [safe_float(row[i]) for i in (2, 3, 4, 5, 6)]
            if any(v is not None for v in env_values):
                report_area = report_area_for_cp(current_cp, ref, fallback_zone)
                main_zone = main_zone_for(current_cp, report_area, ref, fallback_zone)
                env_rows.append(
                    {
                        "Checking Point": display_cp(current_cp),
                        "Report Area": report_area,
                        "Main Zone": main_zone,
                        "Temp. (°C)": maybe_int(env_values[0]),
                        "Humidity (%)": maybe_int(env_values[1]),
                        "CO2 (ppm)": maybe_int(env_values[2]),
                        "Wind Flow (m/s)": maybe_int(clamp_floor(env_values[3], WIND_FLOW_FLOOR)),
                        "PM10 (μg/m³)": maybe_int(env_values[4]),
                    }
                )

            moisture_object = normalize_moisture_object(row[7], row[8])
            moisture_values = [maybe_int(clamp_floor(safe_float(v), MOISTURE_FLOOR)) for v in row[9:14]]
            if moisture_object and any(v is not None for v in moisture_values):
                report_area = report_area_for_cp(current_cp, ref, fallback_zone)
                main_zone = main_zone_for(current_cp, report_area, ref, fallback_zone)
                moisture_rows.append(
                    {
                        "Checking Point": display_cp(current_cp),
                        "Report Area": report_area,
                        "Main Zone": main_zone,
                        "Classification": classify_object(moisture_object, "moisture", ref),
                        "Object": moisture_object,
                        "Moisture (%) - 1": moisture_values[0],
                        "Moisture (%) - 2": moisture_values[1],
                        "Moisture (%) - 3": moisture_values[2],
                        "Moisture (%) - 4": moisture_values[3],
                        "Moisture (%) - 5": moisture_values[4],
                    }
                )

            swab = row[15]
            swab_object = normalize_swab_object(row[16], row[17])
            if swab is not None and swab_object:
                swab_key = normalize_swab_key(swab)
                count = ref.count_by_swab.get(swab_key)
                cfu = count / 0.0003 if count is not None else None
                report_area = report_area_for_cp(current_cp, ref, fallback_zone)
                main_zone = main_zone_for(current_cp, report_area, ref, fallback_zone)
                cfu_rows.append(
                    {
                        "Checking Point": display_cp(current_cp),
                        "Swab": display_cp(swab_key),
                        "Report Area": report_area,
                        "Main Zone": main_zone,
                        "Classification": classify_object(swab_object, "cfu", ref),
                        "Object": swab_object,
                        "Count": maybe_int(count),
                        "CFU/m²": maybe_int(cfu),
                        "Isolate Mold Species": ref.mold_species_by_swab.get(swab_key, []),
                    }
                )

    for outside_name in outside_sheets(wb.sheetnames):
        ws = wb[outside_name]
        current_cp_outside: Any = None
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] is not None:
                current_cp_outside = parse_outside_cp(row[0])
            env_values = [safe_float(row[i]) for i in (1, 2, 3, 4, 5)]
            if not any(v is not None for v in env_values):
                continue
            env_rows.append(
                {
                    "Checking Point": current_cp_outside,
                    "Report Area": "Outside",
                    "Main Zone": "OUTSIDE",
                    "Temp. (°C)": maybe_int(env_values[0]),
                    "Humidity (%)": maybe_int(env_values[1]),
                    "CO2 (ppm)": maybe_int(env_values[2]),
                    "Wind Flow (m/s)": maybe_int(clamp_floor(env_values[3], WIND_FLOW_FLOOR)),
                    "PM10 (μg/m³)": maybe_int(env_values[4]),
                }
            )

    return {"Env. Data": env_rows, "Moisture": moisture_rows, "CFU": cfu_rows}


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for cell in column_cells[:200]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 28)


def write_workbook(data: dict[str, list[dict[str, Any]]], ref: ReferenceData, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, headers in [
        ("Env. Data", ENV_HEADERS),
        ("Moisture", MOISTURE_HEADERS),
        ("CFU", CFU_HEADERS),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in data[sheet_name]:
            ws.append([row.get(h) for h in headers])
        style_sheet(ws)

    ws = wb.create_sheet("Areas")
    ws.append(["Checking Point", "Report Area", None, "Report Area", "Function", "RMW/PL/FGW/OUTSIDE"])
    for cp, area in ref.area_by_cp.items():
        ws.append([display_cp(cp), area, None, None, None, None])
    for cp in ["O1", "O2", "O3", "O4", "O5"]:
        ws.append([cp, "Outside", None, None, None, None])
    row_idx = 2
    for area, info in ref.area_info.items():
        ws.cell(row_idx, 4).value = area
        ws.cell(row_idx, 5).value = info.get("function")
        ws.cell(row_idx, 6).value = info.get("zone")
        row_idx += 1
    ws.cell(row_idx, 4).value = "Outside"
    ws.cell(row_idx, 5).value = "Outdoor reference"
    ws.cell(row_idx, 6).value = "OUTSIDE"
    style_sheet(ws)

    ws = wb.create_sheet("QA Notes")
    ws.append(["Item", "Value"])
    ws.append(["Source", "Generated by smart_ea_automation.py"])
    ws.append(["CFU formula", "CFU/m² = Count / 0.0003"])
    ws.append(["Outside handling", "Outside O1-O5 are appended because they were absent from the manual PYS Env. Data sheet."])
    style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def copy_row_layout(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)
        if isinstance(source.value, str) and source.value.startswith("="):
            target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
        else:
            target.value = source.value


def ensure_min_rows(ws, required_max_row: int, template_row: int | None = None) -> None:
    if required_max_row <= ws.max_row:
        return
    source_row = template_row or ws.max_row
    while ws.max_row < required_max_row:
        copy_row_layout(ws, source_row, ws.max_row + 1)


def clear_columns(ws, start_row: int, end_row: int, columns: list[int]) -> None:
    if end_row < start_row:
        return
    for row_idx in range(start_row, end_row + 1):
        for col_idx in columns:
            ws.cell(row_idx, col_idx).value = None


def fill_formula_columns(ws, start_row: int, end_row: int, columns: list[int], source_row: int = 2) -> None:
    for col_idx in columns:
        source = ws.cell(source_row, col_idx)
        if not (isinstance(source.value, str) and source.value.startswith("=")):
            continue
        for row_idx in range(start_row, end_row + 1):
            target = ws.cell(row_idx, col_idx)
            if row_idx == source_row:
                continue
            target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)


def write_rows_to_template(ws, rows: list[dict[str, Any]], writable_columns: dict[int, str], formula_columns: list[int] | None = None) -> None:
    required_max_row = len(rows) + 1
    ensure_min_rows(ws, required_max_row, template_row=min(ws.max_row, 2))
    clear_columns(ws, 2, ws.max_row, list(writable_columns))
    if formula_columns:
        fill_formula_columns(ws, 2, required_max_row, formula_columns)
    for row_offset, row in enumerate(rows, start=2):
        for col_idx, header in writable_columns.items():
            ws.cell(row_offset, col_idx).value = row.get(header)


def write_template_filled_workbook(data: dict[str, list[dict[str, Any]]], ref: ReferenceData, template_path: Path, output_path: Path) -> None:
    wb = load_workbook(template_path, data_only=False, read_only=False)

    if "Env. Data" in wb.sheetnames:
        write_rows_to_template(
            wb["Env. Data"],
            data["Env. Data"],
            {
                1: "Checking Point",
                4: "Temp. (°C)",
                5: "Humidity (%)",
                6: "CO2 (ppm)",
                7: "Wind Flow (m/s)",
                8: "PM10 (μg/m³)",
            },
            formula_columns=[2, 3],
        )

    if "Moisture" in wb.sheetnames:
        write_rows_to_template(
            wb["Moisture"],
            data["Moisture"],
            {
                1: "Checking Point",
                4: "Classification",
                5: "Object",
                6: "Moisture (%) - 1",
                7: "Moisture (%) - 2",
                8: "Moisture (%) - 3",
                9: "Moisture (%) - 4",
                10: "Moisture (%) - 5",
            },
            formula_columns=[2, 3],
        )

    if "CFU" in wb.sheetnames:
        write_rows_to_template(
            wb["CFU"],
            data["CFU"],
            {
                1: "Checking Point",
                2: "Swab",
                5: "Classification",
                6: "Object",
                7: "Count",
                8: "CFU/m²",
            },
            formula_columns=[3, 4],
        )

    if "Areas" in wb.sheetnames:
        ws = wb["Areas"]
        cp_to_area = {cp_key(row["Checking Point"]): clean_text(row["Report Area"]) for row in data["Env. Data"] if row.get("Checking Point") and row.get("Report Area")}
        area_to_zone: dict[str, str] = {}
        for row in data["Env. Data"]:
            area = clean_text(row.get("Report Area"))
            zone = clean_text(row.get("Main Zone"))
            if area and zone:
                area_to_zone.setdefault(area, zone)
        area_rows: list[tuple[str, str, str]] = [(area, info.get("function", ""), info.get("zone", "")) for area, info in ref.area_info.items()]
        for area, zone in area_to_zone.items():
            if area not in ref.area_info:
                area_rows.append((area, "Outdoor reference" if zone == "OUTSIDE" else "", zone))

        required_rows = max(ws.max_row, len(cp_to_area) + 1, len(area_rows) + 1)
        ensure_min_rows(ws, required_rows, template_row=min(ws.max_row, 2))
        clear_columns(ws, 2, ws.max_row, [1, 2, 4, 5, 6])
        for row_idx, (cp, area) in enumerate(cp_to_area.items(), start=2):
            ws.cell(row_idx, 1).value = display_cp(cp)
            ws.cell(row_idx, 2).value = area

        for row_idx, (area, function, zone) in enumerate(area_rows, start=2):
            ws.cell(row_idx, 4).value = area
            ws.cell(row_idx, 5).value = function
            ws.cell(row_idx, 6).value = zone
        fill_formula_columns(ws, 2, len(area_rows) + 1, [7, 8])

    if "KEY系統" in wb.sheetnames:
        ws = wb["KEY系統"]
        ws["A2"] = '=IF($B$2="","",IFERROR(VLOOKUP($B$2,Areas!$D:$F,3,0),""))'
        ws["C2"] = '=IF($B$2="","",IFERROR(VLOOKUP($B$2,Areas!$D:$E,2,0),""))'

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        pass

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def row_key(row: dict[str, Any], fields: list[str]) -> tuple[str, ...]:
    return tuple(cp_key(row.get(f)) if f in {"Checking Point", "Swab"} else clean_text(row.get(f)) for f in fields)


def values_close(a: Any, b: Any, tolerance: float = 0.011) -> bool:
    af, bf = safe_float(a), safe_float(b)
    if af is not None and bf is not None:
        return abs(af - bf) <= tolerance
    return clean_text(a) == clean_text(b)


def compare_tables(generated: list[dict[str, Any]], target: list[dict[str, Any]], key_fields: list[str], compare_fields: list[str]) -> dict[str, Any]:
    gen_map = {row_key(r, key_fields): r for r in generated}
    target_map = {row_key(r, key_fields): r for r in target}
    missing = [k for k in target_map if k not in gen_map]
    extra = [k for k in gen_map if k not in target_map]
    diffs: list[dict[str, Any]] = []
    for key in sorted(set(gen_map) & set(target_map)):
        for field in compare_fields:
            if not values_close(gen_map[key].get(field), target_map[key].get(field)):
                diffs.append(
                    {
                        "key": key,
                        "field": field,
                        "generated": gen_map[key].get(field),
                        "target": target_map[key].get(field),
                    }
                )
    return {"missing": missing, "extra": extra, "diffs": diffs}


def write_validation_report(data: dict[str, list[dict[str, Any]]], ref: ReferenceData, output_path: Path) -> dict[str, Any]:
    target = ref.target_tables
    results: dict[str, Any] = {}
    if "Env. Data" in target:
        indoor_env = [r for r in data["Env. Data"] if r.get("Main Zone") != "OUTSIDE"]
        results["Env. Data"] = compare_tables(
            indoor_env,
            target["Env. Data"],
            ["Checking Point"],
            ["Report Area", "Main Zone", "Temp. (°C)", "Humidity (%)", "CO2 (ppm)", "Wind Flow (m/s)", "PM10 (μg/m³)"],
        )
    if "Moisture" in target:
        results["Moisture"] = compare_tables(
            data["Moisture"],
            target["Moisture"],
            ["Checking Point", "Object"],
            ["Report Area", "Main Zone", "Classification", "Moisture (%) - 1", "Moisture (%) - 2", "Moisture (%) - 3"],
        )
    if "CFU" in target:
        results["CFU"] = compare_tables(
            data["CFU"],
            target["CFU"],
            ["Checking Point", "Swab"],
            ["Report Area", "Main Zone", "Classification", "Object", "Count", "CFU/m²"],
        )

    outside_rows = [r for r in data["Env. Data"] if r.get("Main Zone") == "OUTSIDE"]
    lines = [
        "# PYS Smart EA Conversion Validation",
        "",
        "## Summary",
        "",
        f"- Generated Env. Data rows: {len(data['Env. Data'])} ({len(outside_rows)} Outside rows included)",
        f"- Generated Moisture rows: {len(data['Moisture'])}",
        f"- Generated CFU rows: {len(data['CFU'])}",
        "- Reference manual workbook was used for Areas, existing object classifications, and CFU colony counts.",
        "",
        "## Table Comparison",
        "",
    ]
    for sheet, result in results.items():
        lines.extend(
            [
                f"### {sheet}",
                "",
                f"- Missing rows vs manual target: {len(result['missing'])}",
                f"- Extra rows vs manual target: {len(result['extra'])}",
                f"- Cell differences: {len(result['diffs'])}",
                "",
            ]
        )
        if result["diffs"]:
            lines.append("| Key | Field | Generated | Manual target |")
            lines.append("|---|---|---:|---:|")
            for diff in result["diffs"][:30]:
                lines.append(f"| {diff['key']} | {diff['field']} | {diff['generated']} | {diff['target']} |")
            if len(result["diffs"]) > 30:
                lines.append(f"| ... | ... | {len(result['diffs']) - 30} more differences omitted | |")
            lines.append("")
        if result["extra"]:
            lines.append("Extra row keys:")
            for key in result["extra"][:20]:
                lines.append(f"- {key}")
            lines.append("")

    lines.extend(
        [
            "## Notes",
            "",
            "- The manual target rounded several low wind-flow values upward (for example 0.05 or 0.08 to 0.1). The generated workbook keeps the source values unless Excel input already contains rounded values.",
            "- Outside O1-O5 rows are expected extras because the manual PYS target currently omits them.",
            "",
        ]
    )
    output_path.write_text("\n".join(lines), encoding="utf-8")
    return results


def numeric_values(rows: list[dict[str, Any]], field: str) -> list[float]:
    values = [safe_float(row.get(field)) for row in rows]
    return [v for v in values if v is not None]


def average(values: list[float]) -> float | None:
    return mean(values) if values else None


def group_area_values(rows: list[dict[str, Any]], field: str) -> dict[str, float]:
    grouped: dict[str, list[float]] = defaultdict(list)
    for row in rows:
        value = safe_float(row.get(field))
        area = clean_text(row.get("Report Area"))
        if area and value is not None:
            grouped[area].append(value)
    return {area: mean(values) for area, values in grouped.items() if values}


def moisture_average(row: dict[str, Any]) -> float | None:
    values = [safe_float(row.get(f"Moisture (%) - {i}")) for i in range(1, 6)]
    values = [v for v in values if v is not None]
    return mean(values) if values else None


def class_label(value: Any) -> str:
    text = clean_text(value)
    return CLASS_LABELS.get(text, text or "unclassified objects")


def backend_class_label(value: Any) -> str:
    text = clean_text(value)
    if text in BACKEND_CLASS_LABELS:
        return BACKEND_CLASS_LABELS[text]
    if "." in text:
        text = text.split(".", 1)[1]
    return "其它" if text == "其他" else text


def sort_area_key(value: Any) -> tuple[int, Any]:
    text = clean_text(value)
    number = safe_float(text)
    if number is not None:
        return (0, number)
    match = re.fullmatch(r"Area\s+(\d+)", text, flags=re.IGNORECASE)
    if match:
        return (0, int(match.group(1)))
    return (1, text)


def compact_values(row: dict[str, Any], fields: list[str]) -> list[float | int]:
    values: list[float | int] = []
    for field in fields:
        value = safe_float(row.get(field))
        if value is not None:
            compacted = maybe_int(value)
            if compacted is not None:
                values.append(compacted)
    return values


def normalize_row_location(row: dict[str, Any], ref: ReferenceData, fallback_zone: str = "") -> None:
    """Normalize area/zone fields for rows copied from target workbook tables.

    Vendor-derived rows already go through area_by_cp. Rows copied directly
    from target tables, especially CFU, may still contain pure-number report
    areas such as "1". If left as-is, the backend payload creates both
    "1" and "Area 1", which puts CFU into a duplicate area.
    """
    cp = cp_key(row.get("Checking Point"))
    report_area = normalize_area_name(row.get("Report Area"))
    if not report_area and cp:
        report_area = report_area_for_cp(cp, ref, fallback_zone)
    row["Report Area"] = report_area

    main_zone = clean_text(row.get("Main Zone"))
    if not main_zone and report_area:
        main_zone = main_zone_for(cp, report_area, ref, fallback_zone)
    row["Main Zone"] = main_zone


def append_area_payload(area_map: dict[str, dict[str, Any]], area: str, zone: str, ref: ReferenceData) -> dict[str, Any]:
    if area not in area_map:
        info = ref.area_info.get(area, {})
        area_map[area] = {
            "assessment_area": area,
            "function": info.get("function", ""),
            "main_zone": info.get("zone") or zone,
            "checking_points": [],
            "env": {backend_field: [] for backend_field in ENV_BACKEND_FIELDS.values()},
            "moisture": [],
            "microbiology": [],
        }
    return area_map[area]


def build_backend_payload(data: dict[str, list[dict[str, Any]]], ref: ReferenceData, case_name: str) -> dict[str, Any]:
    indoor_area_map: dict[str, dict[str, Any]] = {}
    outside_area_map: dict[str, dict[str, Any]] = {}

    for row in data["Env. Data"]:
        zone = clean_text(row.get("Main Zone"))
        area = normalize_area_name(row.get("Report Area"))
        cp = row.get("Checking Point")
        if not area:
            continue
        target_map = outside_area_map if zone == "OUTSIDE" else indoor_area_map
        area_payload = append_area_payload(target_map, area, zone, ref)
        if cp not in area_payload["checking_points"]:
            area_payload["checking_points"].append(cp)
        for source_field, backend_field in ENV_BACKEND_FIELDS.items():
            value = safe_float(row.get(source_field))
            if value is not None:
                area_payload["env"][backend_field].append(maybe_int(value))

    for row in data["Moisture"]:
        zone = clean_text(row.get("Main Zone"))
        area = normalize_area_name(row.get("Report Area"))
        if not area:
            continue
        target_map = outside_area_map if zone == "OUTSIDE" else indoor_area_map
        area_payload = append_area_payload(target_map, area, zone, ref)
        values = compact_values(row, [f"Moisture (%) - {idx}" for idx in range(1, 6)])
        if values:
            area_payload["moisture"].append(
                {
                    "checking_point": row.get("Checking Point"),
                    "classification_source": clean_text(row.get("Classification")),
                    "category_to_select": backend_class_label(row.get("Classification")),
                    "object": clean_text(row.get("Object")),
                    "values": values,
                }
            )

    for row in data["CFU"]:
        zone = clean_text(row.get("Main Zone"))
        area = normalize_area_name(row.get("Report Area"))
        if not area:
            continue
        target_map = outside_area_map if zone == "OUTSIDE" else indoor_area_map
        area_payload = append_area_payload(target_map, area, zone, ref)
        count = safe_float(row.get("Count"))
        area_payload["microbiology"].append(
            {
                "checking_point": row.get("Checking Point"),
                "swab": row.get("Swab"),
                "classification_source": clean_text(row.get("Classification")),
                "category_to_select": backend_class_label(row.get("Classification")),
                "object": clean_text(row.get("Object")),
                "count_to_enter": maybe_int(count),
                "cfu_per_m2_for_check": maybe_int(safe_float(row.get("CFU/m²"))),
                "isolate_mold_species_names": row.get("Isolate Mold Species", []),
            }
        )

    def grouped_areas(area_map: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
        by_zone: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for area in area_map.values():
            by_zone[clean_text(area.get("main_zone"))].append(area)
        groups = []
        zone_order = {"RMW": 1, "PL": 2, "FGW": 3, "OUTSIDE": 4}
        for zone in sorted(by_zone, key=lambda z: (zone_order.get(z, 99), z)):
            areas = sorted(by_zone[zone], key=lambda a: sort_area_key(a["assessment_area"]))
            groups.append(
                {
                    "main_zone": zone,
                    "main_zone_label": ZONE_LABELS.get(zone, zone),
                    "areas": areas,
                }
            )
        return groups

    quality_issues: list[dict[str, Any]] = []
    for table_name in ["Env. Data", "Moisture", "CFU"]:
        for row in data[table_name]:
            if not clean_text(row.get("Report Area")):
                quality_issues.append(
                    {
                        "table": table_name,
                        "checking_point": row.get("Checking Point"),
                        "swab": row.get("Swab"),
                        "issue": "missing_report_area",
                    }
                )
    for row in data["CFU"]:
        if safe_float(row.get("Count")) is None:
            quality_issues.append(
                {
                    "table": "CFU",
                    "checking_point": row.get("Checking Point"),
                    "swab": row.get("Swab"),
                    "issue": "missing_count",
                }
            )

    return {
        "case_name": case_name,
        "purpose": "backend_input_payload",
        "backend_rules": {
            "environment": "Fill exactly the listed values for each field. Remove or avoid unused blank input boxes because the backend average may count blanks as zero.",
            "moisture": "Select category_to_select, enter object, then enter all values in values.",
            "microbiology": "Enter count_to_enter into the backend 數值 field. cfu_per_m2_for_check is only for manual verification.",
        },
        "indoor": grouped_areas(indoor_area_map),
        "outside": grouped_areas(outside_area_map),
        "quality_issues": quality_issues,
        "summary": {
            "indoor_area_count": len(indoor_area_map),
            "outside_area_count": len(outside_area_map),
            "env_rows": len(data["Env. Data"]),
            "moisture_rows": len(data["Moisture"]),
            "microbiology_rows": len(data["CFU"]),
            "quality_issue_count": len(quality_issues),
        },
    }


def write_backend_payload(data: dict[str, list[dict[str, Any]]], ref: ReferenceData, case_name: str, output_path: Path) -> dict[str, Any]:
    payload = build_backend_payload(data, ref, case_name)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return payload


def run_automation(vendor_path: Path | str, target_path: Path | str | None, outdir: Path | str, case_name: str) -> dict[str, Any]:
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    vendor_path = Path(vendor_path)
    target_path = Path(target_path) if target_path else None
    ref = read_reference(target_path)
    data = read_vendor(vendor_path, ref)

    # 公式範本 CFU sheet 是採樣計畫主檔（Checking Point/Swab/Object 由業務/實驗端填）
    # Count 留空時從「菌落計數表」算好的 count_by_swab 回填，避免 CFU sheet G 欄被清空
    manual_cfu = ref.target_tables.get("CFU", [])
    for r in manual_cfu:
        normalize_row_location(r, ref)
        swab_key = normalize_swab_key(r.get("Swab"))
        count = safe_float(r.get("Count"))
        if count is None:
            count = ref.count_by_swab.get(swab_key)
            if count is not None:
                r["Count"] = maybe_int(count)
        if count is not None and safe_float(r.get("CFU/m²")) is None:
            r["CFU/m²"] = maybe_int(count / 0.0003)
    if manual_cfu and any(r.get("Checking Point") is not None for r in manual_cfu):
        data["CFU"] = [r for r in manual_cfu if r.get("Checking Point") is not None]
    # Fill missing Classification for CFU rows using object name heuristics
    for row in data["CFU"]:
        if not clean_text(row.get("Classification")):
            obj = clean_text(row.get("Object", ""))
            if obj:
                row["Classification"] = classify_object(obj, "cfu", ref)
        row["Isolate Mold Species"] = ref.mold_species_by_swab.get(normalize_swab_key(row.get("Swab")), [])

    workbook_path = outdir / f"{case_name}_standardized_from_vendor.xlsx"
    formula_workbook_path = outdir / f"{case_name}_template_filled_with_formulas.xlsx"
    validation_path = outdir / f"{case_name}_conversion_validation.md"
    backend_payload_path = outdir / f"{case_name}_backend_input_payload.json"
    yims_fill_plan_path = outdir / f"{case_name}_yims_fill_plan.json"
    yims_fill_plan_md_path = outdir / f"{case_name}_yims_fill_plan.md"

    write_workbook(data, ref, workbook_path)
    if target_path and target_path.exists():
        write_template_filled_workbook(data, ref, target_path, formula_workbook_path)
    validation = write_validation_report(data, ref, validation_path)
    backend_payload = write_backend_payload(data, ref, case_name, backend_payload_path)
    write_yims_fill_plan(backend_payload, yims_fill_plan_path, yims_fill_plan_md_path)

    paths = {
        "standardized_workbook": workbook_path,
        "formula_workbook": formula_workbook_path if target_path and target_path.exists() else None,
        "validation_report": validation_path,
        "backend_payload": backend_payload_path,
        "yims_fill_plan_json": yims_fill_plan_path,
        "yims_fill_plan_md": yims_fill_plan_md_path,
    }
    return {
        "case_name": case_name,
        "outdir": outdir,
        "paths": paths,
        "validation": validation,
        "cfu_data": data["CFU"],
        "backend_summary": backend_payload.get("summary", {}),
        "quality_issues": backend_payload.get("quality_issues", []),
    }


def run(args: argparse.Namespace) -> None:
    result = run_automation(args.vendor, args.target, args.outdir, args.case_name)
    paths = result["paths"]

    print(f"Generated workbook: {paths['standardized_workbook']}")
    if paths["formula_workbook"]:
        print(f"Formula-preserved workbook: {paths['formula_workbook']}")
    print(f"Validation report: {paths['validation_report']}")
    print(f"Backend input payload: {paths['backend_payload']}")
    print(f"YIMS fill plan JSON: {paths['yims_fill_plan_json']}")
    print(f"YIMS fill plan Markdown: {paths['yims_fill_plan_md']}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Smart EA report automation")
    parser.add_argument("--vendor", default=str(DEFAULT_VENDOR), help="Vendor field-data Excel path")
    parser.add_argument("--target", default=str(DEFAULT_TARGET), help="Reference standardized Excel path")
    parser.add_argument("--outdir", default=str(DEFAULT_OUTDIR), help="Output directory")
    parser.add_argument("--case-name", default="PYS", help="Case prefix for generated files")
    return parser


if __name__ == "__main__":
    run(build_parser().parse_args())
