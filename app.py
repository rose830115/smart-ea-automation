from __future__ import annotations

import json
import re
import os
import subprocess
import sys
import time
import warnings
from pathlib import Path
from typing import Any

import streamlit as st
from openpyxl import load_workbook

from smart_ea_automation import run_automation, run_from_standardized
from validation import validate_inputs


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_ROOT = PROJECT_DIR / "web_outputs"
UPLOAD_ROOT = PROJECT_DIR / "uploads"
COMMENT_GENERATOR_VERSION = "cross-zone-overall-20260513-v3"

IS_CLOUD = str(PROJECT_DIR).startswith("/mount/src")

VENDOR_SHEETS = {"Raw Material warehouse", "Production Line", "Finished Goods Warehouse"}
TARGET_HINT_SHEETS = {"Areas", "Env. Data", "Moisture", "CFU"}

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")


def sanitize_case_name(value: str) -> str:
    value = value.strip() or "Smart_EA_case"
    value = re.sub(r"[^\w.-]+", "_", value, flags=re.UNICODE)
    value = value.strip("._")
    return value or "Smart_EA_case"


def workbook_sheets(path: Path) -> set[str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        return set(wb.sheetnames)
    except Exception:
        return set()


def excel_files_in(case_dir: Path) -> list[Path]:
    if not case_dir.exists() or not case_dir.is_dir():
        return []
    return sorted(
        [
            path
            for path in case_dir.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~$") and not path.name.startswith(".")
        ],
        key=lambda p: p.name.casefold(),
    )


def detect_workbooks(case_dir: Path) -> tuple[list[Path], list[Path]]:
    vendor_candidates: list[Path] = []
    target_candidates: list[Path] = []

    for path in excel_files_in(case_dir):
        sheets = workbook_sheets(path)
        if VENDOR_SHEETS.issubset(sheets):
            vendor_candidates.append(path)
        if TARGET_HINT_SHEETS.issubset(sheets) or {"Areas", "ITEMLIST"}.issubset(sheets):
            target_candidates.append(path)

    return vendor_candidates, target_candidates


def choose_default(options: list[Path], preferred_words: tuple[str, ...] = ()) -> int:
    if not options:
        return 0
    for idx, path in enumerate(options):
        name = path.name.casefold()
        if any(word.casefold() in name for word in preferred_words):
            return idx
    return 0


def path_label(path: Path) -> str:
    return str(path)


def save_uploaded_file(uploaded_file: Any, target_dir: Path) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / uploaded_file.name
    target_path.write_bytes(uploaded_file.getbuffer())
    return target_path


def show_download(path: Path, label: str) -> None:
    if not path or not path.exists():
        return
    st.download_button(
        label=label,
        data=path.read_bytes(),
        file_name=path.name,
        mime="application/octet-stream",
        use_container_width=True,
    )


def run_case(case_name: str, vendor_path: Path, target_path: Path | None, outdir: Path) -> dict[str, Any]:
    return run_automation(
        vendor_path=vendor_path,
        target_path=target_path,
        outdir=outdir,
        case_name=case_name,
    )


def yims_login_ready(account: str = "", password: str = "") -> bool:
    if st.session_state.get("yims_browser_token"):
        return True
    if "YIMS_ACCOUNT" in st.secrets and "YIMS_PASSWORD" in st.secrets:
        return True
    return bool(account.strip() and password)


@st.cache_resource(show_spinner=False)
def install_chromium() -> None:
    subprocess.run(
        [sys.executable, "-m", "playwright", "install", "chromium"],
        capture_output=True,
    )


def playwright_yims_login(account: str, password: str) -> tuple[str | None, list, str | None]:
    from yims_playwright_login import login_with_debug
    return login_with_debug(account, password)


def check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True
    st.title("MRC Smart EA 小工具")
    pwd = st.text_input("請輸入工具密碼", type="password", key="password_input")
    if pwd:
        if pwd == st.secrets.get("app_password", ""):
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("密碼錯誤")
    return False


def run_yims_bot(
    order_id: str,
    paths: dict[str, Path],
    outdir: Path,
    save: bool,
    account: str = "",
    password: str = "",
) -> dict[str, Any]:
    logs_dir = outdir / "99_logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    screenshot_path = logs_dir / f"yims_{order_id}_{'saved' if save else 'preview'}.png"
    risk_json_path = outdir / f"yims_{order_id}_risk_data.json"
    cmd = [
        sys.executable,
        str(PROJECT_DIR / "yims_api_client.py"),
        "--payload",
        str(paths["backend_payload"]),
        "--order-id",
        order_id,
        "--outdir",
        str(outdir),
    ]
    if paths.get("metrics_json"):
        cmd.extend(["--metrics", str(paths["metrics_json"])])
    if save:
        cmd.append("--save")

    env = os.environ.copy()
    if st.session_state.get("yims_browser_token"):
        env["YIMS_TOKEN"] = st.session_state["yims_browser_token"]
    elif "YIMS_ACCOUNT" in st.secrets:
        env["YIMS_ACCOUNT"] = str(st.secrets["YIMS_ACCOUNT"])
        if "YIMS_PASSWORD" in st.secrets:
            env["YIMS_PASSWORD"] = str(st.secrets["YIMS_PASSWORD"])
    elif account.strip() and password:
        env["YIMS_ACCOUNT"] = account.strip()
        env["YIMS_PASSWORD"] = password

    started = time.strftime("%Y-%m-%d %H:%M:%S")
    completed = subprocess.run(
        cmd,
        cwd=str(PROJECT_DIR),
        env=env,
        text=True,
        capture_output=True,
        timeout=240,
    )
    return {
        "command": " ".join(cmd),
        "returncode": completed.returncode,
        "stdout": completed.stdout,
        "stderr": completed.stderr,
        "screenshot": screenshot_path,
        "risk_json": risk_json_path,
        "started": started,
        "save": save,
    }


def risk_numbers_available(risk_data: dict[str, Any]) -> bool:
    for key in ("env_risk", "micro_risk", "overall_risk"):
        values = risk_data.get(key)
        if isinstance(values, dict) and any(value is not None for value in values.values()):
            return True
    return False


def risk_section_available(risk_data: dict[str, Any], key: str) -> bool:
    values = risk_data.get(key)
    return isinstance(values, dict) and any(value is not None for value in values.values())


st.set_page_config(page_title="MRC Smart EA Tool", layout="wide")

if not check_password():
    st.stop()

install_chromium()

if st.session_state.get("comment_generator_version") != COMMENT_GENERATOR_VERSION:
    st.session_state.pop("last_generated_comments", None)
    st.session_state.pop("last_comments_path", None)
    st.session_state["comment_generator_version"] = COMMENT_GENERATOR_VERSION

st.title("MRC Smart EA 小工具")
st.caption("案件資料整理、環境評論草稿、後台輸入資料包與後台填表")

with st.sidebar:
    st.header("案件設定")
    case_name_input = st.text_input("案件名稱", value="SEA_test")
    case_name = sanitize_case_name(case_name_input)
    yims_order_id = st.text_input("後台案件代號", value="", help="用於填入後台資料。")

    st.divider()
    st.header("後台登入")
    if st.session_state.get("yims_browser_token"):
        st.success("後台已登入")
        if st.button("登出後台", use_container_width=True):
            st.session_state.pop("yims_browser_token", None)
            st.rerun()
        yims_account = ""
        yims_password = ""
    else:
        yims_account = st.text_input("後台帳號")
        yims_password = st.text_input("後台密碼", type="password")
        if st.button("登入後台", use_container_width=True, disabled=not (yims_account and yims_password)):
            with st.spinner("正在登入後台，請稍候..."):
                token, screenshots, error = playwright_yims_login(yims_account, yims_password)
            if token:
                st.session_state["yims_browser_token"] = token
                st.rerun()
            else:
                st.error(f"後台登入失敗：{error}")
                import base64
                for i, shot_b64 in enumerate(screenshots):
                    st.image(base64.b64decode(shot_b64), caption=f"截圖 {i+1}")

source_ready = False
vendor_path: Path | None = None
target_path: Path | None = None
adjusted_path: Path | None = None
outdir: Path | None = None

MODE_FRESH = "① 從業務 Excel 整理"
MODE_ADJUSTED = "② 用調整後的標準化 Excel 直接建後台資料"
input_mode = st.radio(
    "資料來源",
    [MODE_FRESH, MODE_ADJUSTED],
    horizontal=True,
    help="② 用在：資料統整後你們又手動改了分類，把調整好的標準化 Excel 丟回來，"
    "工具直接用它重建後台資料包，不用一格一格手動輸入後台。",
)

upload_dir = UPLOAD_ROOT / f"{case_name}_{int(time.time())}"
outdir = DEFAULT_OUTPUT_ROOT / case_name

if input_mode == MODE_FRESH:
    col1, col2 = st.columns(2)
    with col1:
        vendor_upload = st.file_uploader("上傳業務回傳環境資料 Excel", type=["xlsx"])
    with col2:
        target_upload = st.file_uploader("上傳人工標準化 / 公式範本 Excel", type=["xlsx"])
    if vendor_upload is not None:
        vendor_path = save_uploaded_file(vendor_upload, upload_dir)
    if target_upload is not None:
        target_path = save_uploaded_file(target_upload, upload_dir)
    source_ready = vendor_path is not None
else:
    adjusted_upload = st.file_uploader(
        "上傳調整後的標準化 Excel（工具先前產出的 SEA_xxx_standardized_from_vendor.xlsx）",
        type=["xlsx"],
    )
    if adjusted_upload is not None:
        adjusted_path = save_uploaded_file(adjusted_upload, upload_dir)
    source_ready = adjusted_path is not None

st.divider()

if input_mode == MODE_FRESH and source_ready and vendor_path and outdir:
    st.subheader("準備執行")
    st.write(f"案件名稱：`{case_name}`")
    st.write(f"輸出位置：`{outdir}`")
    st.caption("這一步只整理資料與產生檔案;後台輸入會在處理完成後,於下方的後台區塊另外執行。")

    violations = validate_inputs(vendor_path, target_path)
    if violations:
        st.error(
            "上傳的檔案未通過事前檢查,請修正下列項目後重新上傳:\n\n"
            + "\n".join(f"- {v}" for v in violations)
        )
        st.caption(
            "規範參考：`200_Reference/templates/smart-ea_業務交檔規格.md` 與 "
            "`smart-ea_公式範本填表規格.md`"
        )
    else:
        st.success("事前檢查通過,可以開始跑流水線。")

    run_button = st.button(
        "開始處理案件",
        type="primary",
        use_container_width=True,
        disabled=bool(violations),
    )
    if run_button:
        try:
            with st.spinner("處理中，正在整理 Excel、產生評論與輸出檔案..."):
                result = run_case(case_name, vendor_path, target_path, outdir)
            st.session_state["last_result"] = result
            st.session_state["last_cfu_data"] = result.get("cfu_data", [])
            st.success("處理完成")
        except Exception as exc:
            st.error("處理失敗，請確認 Excel 格式與資料夾路徑。")
            st.exception(exc)

elif input_mode == MODE_ADJUSTED and source_ready and adjusted_path:
    # 從檔名自動偵測案件名稱（SEA_xxx_standardized_from_vendor.xlsx → SEA_xxx），
    # 讓輸出資料夾與菌種來源對得上，不用逼使用者去側欄改案件名稱。
    STD_SUFFIX = "_standardized_from_vendor.xlsx"
    if adjusted_path.name.endswith(STD_SUFFIX):
        case_name_eff = sanitize_case_name(adjusted_path.name[: -len(STD_SUFFIX)])
    else:
        case_name_eff = case_name
    outdir_eff = DEFAULT_OUTPUT_ROOT / case_name_eff

    st.subheader("準備執行（調整後 Excel → 後台資料）")
    st.write(f"偵測案件名稱：`{case_name_eff}`")
    st.write(f"輸出位置：`{outdir_eff}`")
    st.caption("菌種直接讀這份標準化 Excel 的「Isolate Mold Species」欄，不依賴其他檔案；隔天或換電腦重傳也讀得到。")

    run_button_adj = st.button(
        "用這份 Excel 建立後台資料", type="primary", use_container_width=True
    )
    if run_button_adj:
        try:
            with st.spinner("讀取調整後 Excel、依你改過的分類重建後台資料包..."):
                result = run_from_standardized(adjusted_path, outdir_eff, case_name_eff)
            st.session_state["last_result"] = result
            st.session_state["last_cfu_data"] = result.get("cfu_data", [])
            species_source = result.get("species_source")
            species_matched = result.get("species_matched", 0)
            if species_source == "excel_column":
                species_note = f"菌種讀自 Excel 菌種欄，共 {species_matched} 筆。"
            elif species_source == "payload_fallback":
                species_note = f"這份 Excel 沒有菌種欄（舊版），退回從第一次結果帶回 {species_matched} 筆。"
            else:
                species_note = "注意：這份 Excel 沒有菌種資料，微生物菌種會留空。"
            st.success(f"完成。{species_note}可到下方「後台輸入」檢查 / 儲存。")
        except Exception as exc:
            st.error("處理失敗，請確認這是工具先前產出的標準化 Excel（分頁與欄位需維持原樣）。")
            st.exception(exc)

else:
    if input_mode == MODE_ADJUSTED:
        st.info("請上傳調整後的標準化 Excel（工具先前產出的 SEA_xxx_standardized_from_vendor.xlsx）。")
    else:
        st.info("請先提供業務回傳環境資料 Excel。")

result = st.session_state.get("last_result")
if result:
    paths = result["paths"]
    summary = result["backend_summary"]
    quality_issues = result["quality_issues"]

    st.divider()
    st.subheader("輸出結果")
    metric_cols = st.columns(5)
    metric_cols[0].metric("Env. Data", summary.get("env_rows", 0))
    metric_cols[1].metric("Moisture", summary.get("moisture_rows", 0))
    metric_cols[2].metric("Microbiology", summary.get("microbiology_rows", 0))
    metric_cols[3].metric("室內區域", summary.get("indoor_area_count", 0))
    metric_cols[4].metric("資料問題", summary.get("quality_issue_count", 0))

    if quality_issues:
        st.warning("有資料問題需要人工確認。")
        st.dataframe(quality_issues, use_container_width=True)

    download_cols = st.columns(2)
    with download_cols[0]:
        show_download(paths["standardized_workbook"], "下載整理後 Excel")
    with download_cols[1]:
        show_download(paths["validation_report"], "下載比對報告")

    st.subheader("後台輸入")
    st.info("下方按鈕改用後台 API，不開瀏覽器。檢查只驗證登入、案件、資料包與送出格式；勾選確認後才會寫入後台。")

    if not yims_order_id.strip():
        st.warning("請先在左側輸入後台案件代號，才能執行後台填表。")
    else:
        login_ready = yims_login_ready(yims_account, yims_password)
        if not login_ready:
            st.warning("請在左側輸入後台帳號與密碼。")
        yims_cols = st.columns(2)
        with yims_cols[0]:
            if st.button("檢查後台（不儲存）", use_container_width=True):
                with st.spinner("正在檢查後台、案件資料與送出格式..."):
                    try:
                        yims_result = run_yims_bot(
                            yims_order_id.strip(),
                            paths,
                            result["outdir"],
                            save=False,
                            account=yims_account,
                            password=yims_password,
                        )
                        st.session_state["last_yims_result"] = yims_result
                    except subprocess.TimeoutExpired:
                        st.error("後台填表逾時。請確認網路與案件代號。")
        with yims_cols[1]:
            confirm_save = st.checkbox("我確認要儲存到後台")
            if st.button("快速儲存到後台", type="primary", use_container_width=True, disabled=not confirm_save):
                with st.spinner("正在透過後台 API 儲存..."):
                    try:
                        yims_result = run_yims_bot(
                            yims_order_id.strip(),
                            paths,
                            result["outdir"],
                            save=True,
                            account=yims_account,
                            password=yims_password,
                        )
                        st.session_state["last_yims_result"] = yims_result
                    except subprocess.TimeoutExpired:
                        st.error("後台儲存逾時。請到後台確認是否已儲存。")

    yims_result = st.session_state.get("last_yims_result")
    if yims_result:
        if yims_result["returncode"] == 0:
            st.success("後台動作完成。")
            if yims_result["save"]:
                st.write("這次已透過後台 API 儲存。")
            else:
                st.write("這次只有檢查後台與資料包，沒有儲存到後台。")
        else:
            st.error("後台動作失敗。")
            st.code(yims_result["stderr"] or yims_result["stdout"], language="text")
            st.warning("若錯誤內容提到尚未登入，請先在工具主機終端機手動登入一次後台，或設定 YIMS_ACCOUNT / YIMS_PASSWORD。")


    # -----------------------------------------------------------------------
    # Phase 3: Generate comments using ChatGPT after YIMS risk data is ready
    # -----------------------------------------------------------------------
    yims_result = st.session_state.get("last_yims_result")
    if yims_result and yims_result.get("returncode") == 0:
        st.divider()
        st.subheader("第三步：後台風險確認與評論生成")

        risk_json_path: Path = yims_result.get("risk_json")
        risk_data: dict = {}

        if risk_json_path and risk_json_path.exists():
            try:
                risk_data = json.loads(risk_json_path.read_text(encoding="utf-8"))
                if "error" in risk_data and "env_risk" not in risk_data:
                    st.warning(
                        "✅ 資料已儲存到後台，但下游「風險圖表資料」抓取失敗（後台 500）。\n\n"
                        "常見原因：該案件尚未在後台基本資料設定「分析參數選擇」，或後台前端的圖表狀態異常。"
                        "請手動到後台確認該案件，並在下方手動填入風險值。"
                    )
                    st.code(risk_data["error"], language="text")
                    risk_data = {}
                elif not risk_numbers_available(risk_data):
                    st.warning(
                        "後台沒有回傳發霉風險數字。請確認後台基本資料已設定「分析參數選擇」，"
                        "並且可視性風險資料已由人工確認儲存；確認後再重新執行後台儲存。"
                    )
                elif not risk_section_available(risk_data, "env_risk") or not risk_section_available(risk_data, "micro_risk"):
                    st.warning(
                        "後台風險數字不完整：環境面或微生物面尚未回傳。"
                        "請到後台確認風險圖表已正常顯示後，再重新執行後台儲存。"
                    )
                else:
                    st.success(f"已從後台自動抓取風險數據：`{risk_json_path.name}`")
            except Exception:
                st.warning("風險數據 JSON 讀取失敗，請手動填入下方數值。")
        else:
            st.warning("尚未抓取到後台風險數據（請先執行後台填表）。可手動填入下方數值。")

        # Show report settings screenshot if available
        risk_screenshot = risk_data.get("risk_screenshot")
        if risk_screenshot and Path(risk_screenshot).exists():
            with st.expander("後台報告設定截圖", expanded=True):
                st.image(risk_screenshot, use_container_width=True)

        # Show raw table data if scraped
        raw_tables = risk_data.get("raw_tables", [])
        if raw_tables:
            with st.expander("後台抓取到的原始表格", expanded=False):
                for tbl in raw_tables[:3]:
                    st.dataframe(tbl)

        # Risk number confirmation / manual input
        st.markdown("**確認或修正後台風險數值**（若自動抓取正確可直接跳過）")
        auto_env = risk_data.get("env_risk", {})
        auto_micro = risk_data.get("micro_risk", {})
        auto_overall = risk_data.get("overall_risk", {})

        risk_cols = st.columns(3)
        with risk_cols[0]:
            st.markdown("**環境面 Environmental Mold Risk (%)**")
            env_rmw = st.number_input("RMW", value=float(auto_env.get("RMW") or 0.0), min_value=0.0, max_value=100.0, step=0.1, key="env_rmw")
            env_pl  = st.number_input("PL",  value=float(auto_env.get("PL") or 0.0),  min_value=0.0, max_value=100.0, step=0.1, key="env_pl")
            env_fgw = st.number_input("FGW", value=float(auto_env.get("FGW") or 0.0), min_value=0.0, max_value=100.0, step=0.1, key="env_fgw")
        with risk_cols[1]:
            st.markdown("**微生物面 Microbiology Mold Risk (%)**")
            micro_rmw = st.number_input("RMW", value=float(auto_micro.get("RMW") or 0.0), min_value=0.0, max_value=100.0, step=0.1, key="micro_rmw")
            micro_pl  = st.number_input("PL",  value=float(auto_micro.get("PL") or 0.0),  min_value=0.0, max_value=100.0, step=0.1, key="micro_pl")
            micro_fgw = st.number_input("FGW", value=float(auto_micro.get("FGW") or 0.0), min_value=0.0, max_value=100.0, step=0.1, key="micro_fgw")
        with risk_cols[2]:
            st.markdown("**整體 Overall Mold Risk (%)**")
            overall_rmw = st.number_input("RMW", value=float(auto_overall.get("RMW") or 0.0), min_value=0.0, max_value=100.0, step=0.1, key="overall_rmw")
            overall_pl  = st.number_input("PL",  value=float(auto_overall.get("PL") or 0.0),  min_value=0.0, max_value=100.0, step=0.1, key="overall_pl")
            overall_fgw = st.number_input("FGW", value=float(auto_overall.get("FGW") or 0.0), min_value=0.0, max_value=100.0, step=0.1, key="overall_fgw")

        confirmed_risk_data = {
            "env_risk":   {"RMW": env_rmw or None,   "PL": env_pl or None,   "FGW": env_fgw or None},
            "micro_risk": {"RMW": micro_rmw or None, "PL": micro_pl or None, "FGW": micro_fgw or None},
            "overall_risk": {"RMW": overall_rmw or None, "PL": overall_pl or None, "FGW": overall_fgw or None},
        }

        if st.button("產生評論文字檔", type="primary", use_container_width=True):
            try:
                from rule_based_comment_generator import generate_rule_based_comments, load_json, write_comments_txt

                backend_payload = load_json(paths["backend_payload"])
                generated = generate_rule_based_comments(backend_payload, confirmed_risk_data)
                comments_path = write_comments_txt(generated, result["outdir"], case_name)
                st.session_state["last_generated_comments"] = generated
                st.session_state["last_comments_path"] = comments_path
                st.success("評論文字檔已產生。")
            except Exception as exc:
                st.error("評論文字檔產生失敗。")
                st.exception(exc)

        generated_comments = st.session_state.get("last_generated_comments")
        if generated_comments:
            sections = [
                ("RMW_env", "Raw Material Warehouse — Environmental"),
                ("PL_env", "Production Line — Environmental"),
                ("FGW_env", "Finished Goods Warehouse — Environmental"),
                ("OVERALL_ENV", "Overall Environmental Comment"),
                ("RMW_micro", "Raw Material Warehouse — Microbiology"),
                ("PL_micro", "Production Line — Microbiology"),
                ("FGW_micro", "Finished Goods Warehouse — Microbiology"),
                ("OVERALL_MICRO", "Overall Microbiology Comment"),
                ("OVERALL_COMPREHENSIVE", "Comprehensive Overall Mold Risk Comment"),
            ]
            for key, title in sections:
                with st.expander(title, expanded=(key == "OVERALL_COMPREHENSIVE")):
                    st.write(generated_comments.get(key, "（未生成）"))

            comments_path = st.session_state.get("last_comments_path")
            if comments_path and Path(comments_path).exists():
                path_obj = Path(comments_path)
                download_label = "下載完整評論文字檔" if path_obj.suffix == ".txt" else "下載完整評論 Markdown"
                show_download(path_obj, download_label)
